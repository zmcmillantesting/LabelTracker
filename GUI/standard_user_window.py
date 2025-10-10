# standard_user_window.py - Workflow Layout Version

import logging
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, 
    QTableWidget, QTableWidgetItem, QLabel, QCheckBox, QMessageBox, 
    QTextEdit, QDialog, QDialogButtonBox, QFrame, QScrollArea, QSplitter
)
from PyQt5.QtWidgets import QSizePolicy
from PyQt5.QtWidgets import QHeaderView
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
from datetime import datetime
from openpyxl import load_workbook
import os
import time
import tempfile
import GUI.styles as styles

logger = logging.getLogger(__name__)

# How many rows to try to show by default in split-screen (can be overridden by env LT_VISIBLE_ROWS)
try:
    DEFAULT_VISIBLE_ROWS = int(os.environ.get('LT_VISIBLE_ROWS', '20'))
except Exception:
    DEFAULT_VISIBLE_ROWS = 15


class FailureDialog(QDialog):
    """Popup dialog for entering failure explanation"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Failure Explanation")
        self.setModal(True)
        self.setStyleSheet("""
            QDialog {
                background-color: #0f1419;
            }
            QTextEdit {
                background-color: #2a2a4e;
                border: 2px solid #667eea;
                border-radius: 6px;
                color: #eee;
                padding: 10px;
            }
        """)

        layout = QVBoxLayout()
        self.text_edit = QTextEdit()
        self.text_edit.setPlaceholderText("Enter failure explanation...")
        layout.addWidget(self.text_edit)

        self.button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

        self.setLayout(layout)

    def get_text(self):
        return self.text_edit.toPlainText().strip()


class FixDialog(QDialog):
    """Popup dialog for entering fix explanation"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Fix Explanation")
        self.setModal(True)
        self.setStyleSheet("""
            QDialog {
                background-color: #0f1419;
            }
            QTextEdit {
                background-color: #2a2a4e;
                border: 2px solid #667eea;
                border-radius: 6px;
                color: #eee;
                padding: 10px;
            }
        """)

        layout = QVBoxLayout()
        info_label = QLabel("This board was previously marked as failed.\nPlease explain what was fixed:")
        info_label.setStyleSheet("color: #aaa; margin-bottom: 10px;")
        layout.addWidget(info_label)

        self.text_edit = QTextEdit()
        self.text_edit.setPlaceholderText("Enter fix explanation...")
        layout.addWidget(self.text_edit)

        self.button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)

        self.setLayout(layout)

    def get_text(self):
        return self.text_edit.toPlainText().strip()


class WorkflowStep(QFrame):
    """Visual workflow step indicator"""
    def __init__(self, number, text, parent=None):
        super().__init__(parent)
        self.number = number
        self.text = text
        self.active = False
        self.setup_ui()
    
    def setup_ui(self):
        self.setFrameShape(QFrame.StyledPanel)
        # Allow each step to expand so all steps appear visually equal
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)
        
        # Step number circle
        self.number_label = QLabel(str(self.number))
        self.number_label.setAlignment(Qt.AlignCenter)
        self.number_label.setFixedSize(50, 50)
        self.number_label.setStyleSheet("""
            background-color: #2a2a4e;
            color: #aaa;
            border-radius: 25px;
            font-size: 18pt;
            font-weight: bold;
        """)
        layout.addWidget(self.number_label)
        
        # Step text
        self.text_label = QLabel(self.text)
        self.text_label.setAlignment(Qt.AlignCenter)
        self.text_label.setStyleSheet("color: #aaa; margin-top: 10px;")
        layout.addWidget(self.text_label)
        
        self.update_style()
    
    def set_active(self, active):
        self.active = active
        self.update_style()
    
    def update_style(self):
        if self.active:
            self.setStyleSheet("""
                QFrame {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #667eea, stop:1 #764ba2);
                    border-radius: 10px;
                    padding: 15px;
                }
            """)
            self.number_label.setStyleSheet("""
                background-color: white;
                color: #667eea;
                border-radius: 25px;
                font-size: 18pt;
                font-weight: bold;
            """)
            self.text_label.setStyleSheet("color: white; margin-top: 10px; font-weight: bold;")
        else:
            self.setStyleSheet("""
                QFrame {
                    background-color: #2a2a4e;
                    border-radius: 10px;
                    padding: 15px;
                }
            """)
            self.number_label.setStyleSheet("""
                background-color: #1a1a2e;
                color: #aaa;
                border-radius: 25px;
                font-size: 18pt;
                font-weight: bold;
            """)
            self.text_label.setStyleSheet("color: #aaa; margin-top: 10px;")


class UserWindow(QWidget):
    def __init__(self, username: str, user_id: int, db_manager, xlsx_manager, on_logout=None):
        super().__init__()
        self.username = username
        self.user_id = user_id
        self.db_manager = db_manager
        self.xlsx_manager = xlsx_manager
        self.on_logout = on_logout
        
        self.setWindowTitle("Label Tracker - User")
        self.setMinimumSize(1200, 800)
        
        # Custom dark theme
        self.setStyleSheet("""
            QWidget {
                background-color: #0f1419;
                color: #eee;
                font-family: 'Segoe UI';
                font-size: 12pt;
            }
            QLineEdit {
                padding: 12px;
                background-color: #2a2a4e;
                border: 2px solid #3a3a5e;
                border-radius: 6px;
                color: #eee;
                font-size: 14pt;
            }
            QLineEdit:focus {
                border: 2px solid #667eea;
            }
            QPushButton {
                padding: 12px 25px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #667eea, stop:1 #764ba2);
                border: none;
                border-radius: 6px;
                color: white;
                font-weight: bold;
                font-size: 13pt;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #7b8fef, stop:1 #8b5fb7);
            }
            QPushButton:disabled {
                background-color: #2a2a4e;
                color: #666;
            }
            QCheckBox {
                spacing: 10px;
                font-size: 14pt;
            }
            QCheckBox::indicator {
                width: 25px;
                height: 25px;
                border: 2px solid #3a3a5e;
                border-radius: 4px;
                background-color: #2a2a4e;
            }
            QCheckBox::indicator:checked {
                background-color: #667eea;
                border-color: #667eea;
            }
            QTableWidget {
                background-color: #1a1a2e;
                alternate-background-color: #16213e;
                gridline-color: #2a2a4e;
                border: none;
                border-radius: 8px;
            }
            QTableWidget::item {
                padding: 10px;
            }
            QTableWidget::item:selected {
                background-color: #667eea;
            }
            QHeaderView::section {
                background-color: #2a2a4e;
                padding: 12px;
                border: none;
                font-weight: bold;
                color: #64b5f6;
            }
        """)
        
        # Track current order info
        self.current_order_id = None
        self.current_order_file = None
        self.current_company_name = None
        self.current_board_name = None
        
        # Track fail history per serial number
        self.serial_history = {}
        self.current_serial = None
        
        # Workflow steps
        self.workflow_steps = []
        self.current_step = 0

        logger.info(f"User window initialized for: {username}")
        self.setup_ui()

    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Scroll area for entire content
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("QScrollArea { border: none; }")
        
        # Container with max width
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        
    # Center content horizontally
        center_layout = QHBoxLayout()
        center_layout.addStretch()
        
        # Content area (allow expanding to fill available width)
        content_area = QWidget()
        # allow the content area to expand horizontally so tables can grow
        content_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        content_layout = QVBoxLayout(content_area)
        # reduce margins to make better use of side space
        content_layout.setContentsMargins(12, 12, 12, 12)
        content_layout.setSpacing(20)
        
        # Header
        header_layout = QHBoxLayout()
        title = QLabel("Board Testing Workflow")
        title_font = QFont("Segoe UI", 24, QFont.Bold)
        title.setFont(title_font)
        title.setStyleSheet("color: #64b5f6;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        user_label = QLabel(f"Operator: {self.username}")
        user_label.setStyleSheet("color: #aaa; font-size: 14pt;")
        header_layout.addWidget(user_label)
        
        self.logout_button = QPushButton("Logout")
        self.logout_button.setStyleSheet("""
            QPushButton {
                background: #f44336;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: #da190b;
            }
        """)
        self.logout_button.clicked.connect(self.handle_logout)
        header_layout.addWidget(self.logout_button)
        
        content_layout.addLayout(header_layout)
        
        # Workflow steps
        steps_layout = QHBoxLayout()
        steps_layout.setSpacing(10)
        
        self.step1 = WorkflowStep(1, "Load Order", self)
        self.step2 = WorkflowStep(2, "Scan Board", self)
        self.step3 = WorkflowStep(3, "Record Result", self)
        self.step4 = WorkflowStep(4, "Submit", self)
        
        self.workflow_steps = [self.step1, self.step2, self.step3, self.step4]
        
        steps_layout.addWidget(self.step1)
        
        arrow1 = QLabel("→")
        arrow1.setStyleSheet("color: #667eea; font-size: 30pt; font-weight: bold;")
        arrow1.setAlignment(Qt.AlignCenter)
        steps_layout.addWidget(arrow1)
        
        steps_layout.addWidget(self.step2)
        
        arrow2 = QLabel("→")
        arrow2.setStyleSheet("color: #667eea; font-size: 30pt; font-weight: bold;")
        arrow2.setAlignment(Qt.AlignCenter)
        steps_layout.addWidget(arrow2)
        
        steps_layout.addWidget(self.step3)
        
        arrow3 = QLabel("→")
        arrow3.setStyleSheet("color: #667eea; font-size: 30pt; font-weight: bold;")
        arrow3.setAlignment(Qt.AlignCenter)
        steps_layout.addWidget(arrow3)
        
        steps_layout.addWidget(self.step4)
        
        # place the steps into the centered content area
        content_layout.addLayout(steps_layout)
        
        # Set initial step
        self.update_workflow_step(0)
        
        # Order loading section
        order_frame = QFrame()
        order_frame.setStyleSheet("""
            QFrame {
                background-color: #1a1a2e;
                border-radius: 10px;
                padding: 20px;
            }
        """)
        order_layout = QVBoxLayout(order_frame)
        
        order_label = QLabel("Load Order")
        order_label.setStyleSheet("color: #64b5f6; font-size: 16pt; font-weight: bold;")
        order_layout.addWidget(order_label)
        
        search_layout = QHBoxLayout()
        self.order_input = QLineEdit()
        self.order_input.setPlaceholderText("Enter order number...")
        self.order_input.returnPressed.connect(self.load_order_info)
        search_layout.addWidget(self.order_input)
        
        self.search_button = QPushButton("Load Order")
        self.search_button.clicked.connect(self.load_order_info)
        search_layout.addWidget(self.search_button)
        
        order_layout.addLayout(search_layout)
        
        # Order info display
        info_layout = QHBoxLayout()
        self.company_label = QLabel("Company: ---")
        self.company_label.setStyleSheet("color: #aaa;")
        info_layout.addWidget(self.company_label)
        
        self.board_label = QLabel("Board Type: ---")
        self.board_label.setStyleSheet("color: #aaa;")
        info_layout.addWidget(self.board_label)
        
        self.order_status_label = QLabel("Status: ---")
        self.order_status_label.setStyleSheet("color: #aaa;")
        info_layout.addWidget(self.order_status_label)
        
        info_layout.addStretch()
        order_layout.addLayout(info_layout)
        
        # place the order frame into the centered content area
        content_layout.addWidget(order_frame)

    # Testing section (serial input & pass/fail buttons)
        # in the middle of the page (below Load Order)
        test_frame = QFrame()
        test_frame.setStyleSheet(""" 
            QFrame {
                background-color: #1a1a2e;
                border-radius: 10px;
                padding: 20px;
            }
        """)
        test_layout = QVBoxLayout(test_frame)

        test_label = QLabel("Test Board")
        test_label.setStyleSheet("color: #64b5f6; font-size: 16pt; font-weight: bold;")
        test_layout.addWidget(test_label)

        # Serial input with icon and larger height
        sn_layout = QHBoxLayout()
        sn_layout.setSpacing(10)
        sn_icon = QLabel("🔍")
        sn_icon.setStyleSheet("font-size:18pt; color: #64b5f6; padding-right:6px;")
        sn_layout.addWidget(sn_icon)

        self.sn_input = QLineEdit()
        self.sn_input.setPlaceholderText("Scan or enter serial number...")
        self.sn_input.returnPressed.connect(self.on_sn_entered)
        self.sn_input.setMinimumHeight(54)
        self.sn_input.setStyleSheet(
            "QLineEdit { padding: 12px; border-radius: 10px; background-color: #2a2a4e; color: #eee; font-size: 16pt; }"
        )
        sn_layout.addWidget(self.sn_input)

        test_layout.addLayout(sn_layout)

        # Create large Pass/Fail buttons (full-width feel)
        result_layout = QHBoxLayout()
        result_layout.setSpacing(12)

        self.pass_button = QPushButton("✓ Pass")
        self.pass_button.setStyleSheet("background-color: #4caf50; color: white; font-size: 16pt; font-weight: bold; border-radius: 8px;")
        self.pass_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.pass_button.setFixedHeight(60)
        # When clicked, handle pass in an explicit handler
        self.pass_button.clicked.connect(self.on_pass_clicked)
        # disabled until a serial is selected
        self.pass_button.setEnabled(False)

        self.fail_button = QPushButton("✗ Fail")
        self.fail_button.setStyleSheet("background-color: #f44336; color: white; font-size: 16pt; font-weight: bold; border-radius: 8px;")
        self.fail_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.fail_button.setFixedHeight(60)
        # Open failure dialog on fail button (use validation wrapper)
        self.fail_button.clicked.connect(self.on_fail_clicked)
        self.fail_button.setEnabled(False)

        result_layout.addWidget(self.pass_button)
        result_layout.addWidget(self.fail_button)

        # Keep the old checkboxes (hidden) to preserve toggle_checkboxes logic
        self.pass_checkbox = QCheckBox("Pass")
        self.fail_checkbox = QCheckBox("Fail")
        self.pass_checkbox.setVisible(False)
        self.fail_checkbox.setVisible(False)
        self.pass_checkbox.stateChanged.connect(self.toggle_checkboxes)
        self.fail_checkbox.stateChanged.connect(self.toggle_checkboxes)

        # Hidden submit button used by existing code paths
        self.sn_button = QPushButton("Submit Result")
        self.sn_button.setVisible(False)
        self.sn_button.setEnabled(False)
        self.sn_button.clicked.connect(self.submit_sn)

        test_layout.addLayout(result_layout)


        # Add the testing frame into the content area so scan/submit controls are visible
        content_layout.addWidget(test_frame)

        # We'll use a splitter so the testing area stays visible and the bottom table can be resized
        # Create table_label and self.order_table below, then add to a table container

        # Order table (placed at the bottom)
        table_label = QLabel("Order Details")
        table_label.setStyleSheet("color: #64b5f6; font-size: 16pt; font-weight: bold; margin-top: 10px;")
        content_layout.addWidget(table_label)

        self.order_table = QTableWidget()
        self.order_table.setColumnCount(8)
        self.order_table.setHorizontalHeaderLabels([
            "User", "Company", "Board", "Serial Number", 
            "Pass/Fail", "Timestamp", "Failure Explanation", "Fix Explanation"
        ])
        self.order_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.order_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.order_table.setSelectionMode(QTableWidget.SingleSelection)
        self.order_table.cellClicked.connect(self.select_serial)
        self.order_table.setAlternatingRowColors(True)
        # Allow word-wrapping in cells (helps multi-line failure/fix explanations)
        self.order_table.setWordWrap(True)
        # Make order table columns stretch and set a larger minimum height
        try:
            self.order_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.order_table.horizontalHeader().setStretchLastSection(True)
        except Exception:
            pass
        # Default row height and minimum table height: DEFAULT_VISIBLE_ROWS * 30px + header (~48px)
        self.order_table.verticalHeader().setDefaultSectionSize(30)
        self.order_table.setMinimumHeight(DEFAULT_VISIBLE_ROWS * 30 + 48)
        # Allow the table to expand to fill available space (useful in split-screen)
        self.order_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        content_layout.addWidget(self.order_table)
        

        # finish container layout and add centered container to scroll
        # Add content area with stretch so it expands to fill available horizontal space
        center_layout.addWidget(content_area, 1)
        center_layout.addStretch()
        container_layout.addLayout(center_layout)
        scroll.setWidget(container)
        # Add scroll area to main layout (so centered content scrolls)
        main_layout.addWidget(scroll)

    def update_workflow_step(self, step_index):
        """Update which workflow step is active"""
        self.current_step = step_index
        for i, step in enumerate(self.workflow_steps):
            step.set_active(i == step_index)

    def load_order_info(self):
        """Load order info from database and XLSX file"""
        order_number = self.order_input.text().strip()
        if not order_number:
            QMessageBox.warning(self, "Error", "Please enter an order number.")
            return

        try:
            # Find order in database
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT order_id, company_id, board_id, status, file_path FROM orders WHERE order_number=?",
                    (order_number,)
                )
                order = cursor.fetchone()
            
            if not order:
                QMessageBox.warning(self, "Error", f"Order '{order_number}' not found.")
                logger.warning(f"Order not found: {order_number}")
                return
            
            order_id, company_id, board_id, status, file_path = order
            
            # Get company and board names
            companies = {c[0]: c[1] for c in self.db_manager.get_companies()}
            self.current_company_name = companies.get(company_id, "Unknown")
            
            if board_id:
                boards = self.db_manager.get_boards_by_company(company_id)
                board_dict = {b[0]: b[1] for b in boards}
                self.current_board_name = board_dict.get(board_id, "N/A")
            else:
                self.current_board_name = "N/A"
            
            # Update UI labels
            self.company_label.setText(f"Company: {self.current_company_name}")
            self.board_label.setText(f"Board Type: {self.current_board_name}")
            self.order_status_label.setText(f"Status: {status}")
            
            # Store current order info
            self.current_order_id = order_id
            self.current_order_file = file_path
            
            # Load XLSX file data
            self.load_xlsx_data(file_path)
            
            # Move to next step
            self.update_workflow_step(1)
            
            logger.info(f"Order {order_number} loaded successfully")
            
        except Exception as e:
            logger.error(f"Failed to load order: {e}", exc_info=True)
            QMessageBox.critical(self, "Error", f"Failed to load order:\n{str(e)}")

    def load_xlsx_data(self, file_path):
        """Load data from XLSX file into table"""
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Clear existing data
            self.order_table.setRowCount(0)
            self.serial_history.clear()
            
            user_map = {}
            try:
                with self.db_manager.get_connection() as conn:
                    query = conn.cursor()
                    query.execute("SELECT user_id, username FROM users")
                    user_map ={u_id: name for u_id, name in query.fetchall()}
            except Exception as e:
                logger.warning(f'Could not fetch usernames: {e}')

            company_map = {c[0]: c[1] for c in self.db_manager.get_companies_all(include_archived=True)}

            board_map ={}
            for company_id in company_map:
                try:
                    boards = self.db_manager.get_boards_by_company(company_id)
                    for b_id, b_name, *_ in boards:
                        board_map[b_id] = b_name
                except Exception as e:
                    continue

            # Read data (skip header row)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
                user_id, company_id, board_id, serial_number, pass_fail, timestamp, failure_exp, fix_exp = row

                # Normalize serial to a string and strip whitespace to ensure consistent lookup
                username = user_map.get(user_id, "Unknown")
                company_name = company_map.get(company_id, "Unknown")
                board_name = board_map.get(board_id, "Unknown")
                serial_str = str(serial_number).strip() if serial_number is not None else ""

                self.order_table.insertRow(row_idx)
                self.order_table.setItem(row_idx, 0, QTableWidgetItem(username or "Unknown"))
                self.order_table.setItem(row_idx, 1, QTableWidgetItem(company_name or "Unknown"))
                self.order_table.setItem(row_idx, 2, QTableWidgetItem(board_name or ""))
                self.order_table.setItem(row_idx, 3, QTableWidgetItem(serial_str))
                
                # Color code status
                status_item = QTableWidgetItem(str(pass_fail) if pass_fail else "Pending")
                if pass_fail == "Pass":
                    status_item.setForeground(Qt.green)
                elif pass_fail == "Fail":
                    status_item.setForeground(Qt.red)
                else:
                    status_item.setForeground(Qt.yellow)
                self.order_table.setItem(row_idx, 4, status_item)

                formatted_timestamp = ""
                if timestamp:
                    try:
                        dt = datetime.strptime(str(timestamp), "%Y-%m-%d %H:%M:%S")
                        formatted_timestamp = dt.strftime("%b %d, %Y %I:%M %p")
                    except Exception as e:
                        formatted_timestamp - str(timestamp)
                        logger.info(f"Failed to format timestamp: {e}")
                
                self.order_table.setItem(row_idx, 5, QTableWidgetItem(formatted_timestamp))
                self.order_table.setItem(row_idx, 6, QTableWidgetItem(str(failure_exp) if failure_exp else ""))
                self.order_table.setItem(row_idx, 7, QTableWidgetItem(str(fix_exp) if fix_exp else ""))
                
                # Track if this serial was previously failed and store mapping to table row
                was_failed = pass_fail == "Fail"
                self.serial_history[serial_str] = {
                    "was_failed": was_failed,
                    # excel row (for writing back) and table row (for selecting)
                    "row": row_idx + 2,
                    "table_row": row_idx
                }
            
            logger.info(f"Loaded {self.order_table.rowCount()} serial numbers from XLSX")
            # Adjust row heights to fit content: compute a reasonable default from font metrics
            try:
                fm = self.order_table.fontMetrics()
                # Add padding so text isn't cramped; use 1.4x the font height
                default_h = int(fm.height() * 1.4)
                if default_h < 20:
                    default_h = 20
                self.order_table.verticalHeader().setDefaultSectionSize(default_h)
                # Let Qt compute per-row heights based on content where necessary
                self.order_table.resizeRowsToContents()
            except Exception:
                pass
            
        except Exception as e:
            logger.error(f"Failed to load XLSX data: {e}", exc_info=True)
            QMessageBox.critical(self, "Error", f"Failed to load XLSX data:\n{str(e)}")

    def handle_logout(self):
        try:
            self.close()
            if callable(self.on_logout):
                self.on_logout()
        except Exception as e:
            logger.error(f"Error during logout: {e}", exc_info=True)

    def on_sn_entered(self):
        """When user scans/enters an SN"""
        sn = self.sn_input.text().strip()
        if not sn:
            return
        # Ensure an order is loaded
        if not self.current_order_file:
            QMessageBox.warning(self, "Error", "Please load an order first.")
            return

        # Check against loaded XLSX serials (self.serial_history stores normalized strings)
        norm_sn = sn.strip()
        if norm_sn not in self.serial_history:
            QMessageBox.warning(self, "Not Found", f"Serial number '{sn}' is not present in the currently loaded order.")
            self.sn_input.clear()
            return

        # Select the corresponding table row using stored table_row
        try:
            table_row = self.serial_history[norm_sn].get("table_row")
            if table_row is not None and 0 <= table_row < self.order_table.rowCount():
                self.order_table.selectRow(table_row)
                self.current_serial = norm_sn

                # Move to next step and enable action buttons
                self.update_workflow_step(2)
                self.pass_button.setEnabled(True)
                self.fail_button.setEnabled(True)
                self.pass_button.setFocus()
            else:
                QMessageBox.warning(self, "Error", "Could not locate the serial in the table view.")
        except Exception as e:
            logger.error(f"Error selecting scanned SN: {e}", exc_info=True)
            QMessageBox.warning(self, "Error", "An unexpected error occurred while selecting the scanned serial.")

    def select_serial(self, row, column):
        """Select a serial from the table"""
        serial_item = self.order_table.item(row, 3)
        if serial_item:
            self.current_serial = serial_item.text()
            self.sn_input.setText(self.current_serial)
            self.update_workflow_step(2)
            # Enable action buttons when a serial is selected by clicking
            try:
                self.pass_button.setEnabled(True)
                self.fail_button.setEnabled(True)
                self.pass_button.setFocus()
            except Exception:
                pass
        else:
            self.current_serial = None

    def on_pass_clicked(self):
        """Handle Pass button click: keep checkbox compatibility but use explicit flow."""
        # Ensure the serial exists in the loaded XLSX
        if not self.ensure_sn_is_loaded():
            return

        # Keep compatibility with checkbox-driven logic
        self.pass_checkbox.setChecked(True)
        self.fail_checkbox.setChecked(False)

        # If this serial was previously failed, show fix dialog first
        if self.serial_history.get(self.current_serial, {}).get("was_failed", False):
            # If the fix dialog enabled the hidden submit, submit now
            if getattr(self, 'sn_button', None) and self.sn_button.isEnabled():
                self.submit_sn()
        else:
            # Otherwise submit immediately
            self.submit_sn()

    def on_fail_clicked(self):
        """Handle Fail button click; validate SN then open failure dialog."""
        # Ensure the serial exists in the loaded XLSX
        if not self.ensure_sn_is_loaded():
            return

        # Keep compatibility with checkbox-driven logic
        self.pass_checkbox.setChecked(False)
        self.fail_checkbox.setChecked(True)

        # Open the failure dialog which will call submit_sn on accept
        self.open_failure_dialog()


    def reset_action_buttons(self):
        """Reset hidden and visible action controls to default disabled state"""
        try:
            self.pass_checkbox.setChecked(False)
            self.fail_checkbox.setChecked(False)
            self.sn_button.setEnabled(False)
            # disable visible buttons until user interacts or toggle_checkboxes enables
            self.pass_button.setEnabled(False)
            self.fail_button.setEnabled(False)
        except Exception:
            # If called during init before widgets created, ignore
            pass

    def ensure_sn_is_loaded(self) -> bool:
        """Ensure the current SN (visible in sn_input) exists in the loaded XLSX.

        If `self.current_serial` is already set and present, returns True.
        Otherwise tries to validate using the text in `self.sn_input`. If found,
        selects the corresponding table row and returns True. If not found shows
        a warning dialog and returns False.
        """
        # If a serial is already selected and exists in history, accept it
        if self.current_serial and self.normalize_sn(self.current_serial) in self.serial_history:
            return True

        # Otherwise use the input text
        sn_text = (self.sn_input.text() or "").strip()
        if not sn_text:
            QMessageBox.warning(self, "Error", "Please select or enter a serial number first.")
            return False

        norm_sn = self.normalize_sn(sn_text)
        if norm_sn not in self.serial_history:
            QMessageBox.warning(self, "SN not found in file", f"Serial number '{sn_text}' not found in the currently loaded order.")
            return False

        # select the corresponding table row and set current_serial
        table_row = self.serial_history[norm_sn].get("table_row")
        if table_row is not None:
            try:
                self.order_table.selectRow(table_row)
                self.current_serial = norm_sn
                # enable action buttons
                self.pass_button.setEnabled(True)
                self.fail_button.setEnabled(True)
                return True
            except Exception:
                QMessageBox.warning(self, "Error", "Could not select serial in table view.")
                return False

        QMessageBox.warning(self, "Error", "Serial not found in internal mapping.")
        return False

    def toggle_checkboxes(self, state):
        """Ensure only one checkbox is active"""
        if not self.current_serial:
            QMessageBox.warning(self, "Error", "Please select or enter a serial number first.")
            self.pass_checkbox.setChecked(False)
            self.fail_checkbox.setChecked(False)
            return

        if self.sender() == self.pass_checkbox and state == 2:
            self.fail_checkbox.setChecked(False)
            
            # Check if board was previously failed
            if self.serial_history.get(self.current_serial, {}).get("was_failed", False):
                self.open_fix_dialog()
            else:
                self.sn_button.setEnabled(True)
                self.update_workflow_step(3)
                
        elif self.sender() == self.fail_checkbox and state == 2:
            self.pass_checkbox.setChecked(False)

    def open_failure_dialog(self):
        """Popup for failure explanation"""
        dialog = FailureDialog(self)
        result = dialog.exec_()

        if result == QDialog.Accepted:
            explanation = dialog.get_text()
            if not explanation:
                QMessageBox.warning(self, "Error", "Failure explanation is required.")
                self.fail_checkbox.setChecked(False)
                return

            self.submit_sn(failure_reason=explanation)
        else:
            self.fail_checkbox.setChecked(False)

    def open_fix_dialog(self):
        """Popup for fix explanation"""
        dialog = FixDialog(self)
        result = dialog.exec_()

        if result == QDialog.Accepted:
            explanation = dialog.get_text()
            if not explanation:
                QMessageBox.warning(self, "Error", "Fix explanation is required.")
                self.pass_checkbox.setChecked(False)
                return

            self.fix_explanation = explanation
            self.sn_button.setEnabled(True)
            self.update_workflow_step(3)
        else:
            self.pass_checkbox.setChecked(False)

    def submit_sn(self):
        """Handle submission of a serial number result and update the XLSX file."""
        sn = self.serial_input.text().strip()
        if not sn:
            QMessageBox.warning(self, "Warning", "Please enter a serial number.")
            return

        result = self.pass_fail_dropdown.currentText().strip()
        failure_explanation = self.failure_textbox.toPlainText().strip()
        fix_explanation = self.fix_textbox.toPlainText().strip()

        if result not in ["Pass", "Fail"]:
            QMessageBox.warning(self, "Warning", "Please select Pass or Fail before submitting.")
            return

        try:
            wb = load_workbook(self.xlsx_path)
            ws = wb.active

            # Find serial number row
            sn_col_index = 6  # "Serial Number" column index in Excel (1-based)
            pass_fail_col = 7
            timestamp_col = 8
            fail_exp_col = 9
            fix_exp_col = 10

            found_row = None
            for row in range(2, ws.max_row + 1):
                cell_value = str(ws.cell(row=row, column=sn_col_index).value).strip()
                if cell_value == sn:
                    found_row = row
                    break

            if not found_row:
                QMessageBox.warning(self, "Warning", f"Serial number {sn} not found in file.")
                return

            # 🕒 Create timestamp for pass/fail
            timestamp = datetime.now().strftime("%b %d, %Y %I:%M %p")

            # Write results to workbook
            ws.cell(row=found_row, column=pass_fail_col).value = result
            ws.cell(row=found_row, column=timestamp_col).value = timestamp
            ws.cell(row=found_row, column=fail_exp_col).value = failure_explanation if result == "Fail" else ""
            ws.cell(row=found_row, column=fix_exp_col).value = fix_explanation if result == "Fail" else ""

            # Save changes
            wb.save(self.xlsx_path)

            # Update UI table
            for row in range(self.order_table.rowCount()):
                table_sn = self.order_table.item(row, 3).text()
                if table_sn == sn:
                    self.order_table.setItem(row, 4, QTableWidgetItem(result))
                    self.order_table.setItem(row, 5, QTableWidgetItem(timestamp))
                    self.order_table.setItem(row, 6, QTableWidgetItem(failure_explanation))
                    self.order_table.setItem(row, 7, QTableWidgetItem(fix_explanation))
                    break

            QMessageBox.information(self, "Success", f"Serial {sn} marked as {result}.")

            # Reset input fields
            self.serial_input.clear()
            self.failure_textbox.clear()
            self.fix_textbox.clear()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to update file:\n{e}")
            logger.error(f"Failed to update XLSX for serial {sn}: {e}", exc_info=True)

    def normalize_sn(self, sn: str) -> str:
        """Normalize a serial number string for consistent lookup.

        Current strategy: strip whitespace. We can expand this to uppercase,
        remove punctuation, etc., if needed.
        """
        if sn is None:
            return ""
        try:
            return str(sn).strip()
        except Exception:
            return str(sn)

    def update_xlsx_file(self, serial_number, pass_fail, timestamp, failure_explanation="", fix_explanation=""):
        """Update XLSX file with new data"""
        try:
            wb = load_workbook(self.current_order_file)
            ws = wb.active
            # Normalize serial before lookup into serial_history
            norm_sn = self.normalize_sn(serial_number)
            excel_row = self.serial_history[norm_sn]["row"]
            
            ws.cell(row=excel_row, column=1).value = self.user_id
            ws.cell(row=excel_row, column=5).value = pass_fail
            ws.cell(row=excel_row, column=6).value = timestamp
            
            if failure_explanation:
                ws.cell(row=excel_row, column=7).value = failure_explanation
            
            if fix_explanation:
                ws.cell(row=excel_row, column=8).value = fix_explanation
            
            # Safe save
            temp_dir = os.path.dirname(self.current_order_file) or None
            fd, tmp_path = tempfile.mkstemp(prefix="lt_tmp_", suffix=".xlsx", dir=temp_dir)
            os.close(fd)
            try:
                wb.save(tmp_path)
                max_retries = 3
                for attempt in range(1, max_retries + 1):
                    try:
                        os.replace(tmp_path, self.current_order_file)
                        break
                    except PermissionError:
                        if attempt == max_retries:
                            raise
                        time.sleep(0.5)
                logger.info(f"XLSX file updated for SN: {serial_number}")
            finally:
                try:
                    if os.path.exists(tmp_path):
                        os.remove(tmp_path)
                except Exception:
                    pass
            
        except Exception as e:
            logger.error(f"Failed to update XLSX file: {e}", exc_info=True)
            raise