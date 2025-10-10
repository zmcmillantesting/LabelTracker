import os, logging
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime
from managers.db_manager import DatabaseManager
from openpyxl import load_workbook
logger = logging.getLogger(__name__)

class XLSXManager:
    def __init__(self, db: DatabaseManager):
        self.db = db

    #TODO: update parameters to match user input from UI when implemented
    def _generate_serial_numbers(self, prefix=None, start=1, count=1000):
        """Generate serial numbers. If prefix is None, use default placeholder prefix.
        Serial format: {prefix}{sequence:05d}
        """
        if prefix is None:
            prefix = "CUSTID-ORDNO-"
        return [f"{prefix}{str(i).zfill(5)}" for i in range(start, start + count)]

    def create_order_file(
        self, 
        order_number, 
        created_by, 
        user_id, 
        company_id, 
        pass_fail= None, 
        pass_fail_timestamp=None,
        failure_explanation=None, 
        fix_explanation=None, 
        board_id=None, 
        board_name=None, 
        serial_prefix="CUSTID-ORDNO-", 
        serial_start=1, 
        serial_count=50,
        dest_dir: str = None):
        

        """Create a new XLSX file for an order and register it in the DB."""

        # 1. Get company storage path from DB
        companies = self.db.get_companies()
        company = next((c for c in companies if c[0] == company_id), None)
        if not company:
            raise ValueError(f"Company {company_id} not found")

        company_path = company[2]  # client_path from DB
        target_dir = dest_dir if dest_dir else company_path
        os.makedirs(target_dir, exist_ok=True)

        # 2. Build full file path
        filename = f"{order_number}.xlsx"
        file_path = os.path.join(target_dir, filename)

        # 3. get usernames
        username = "Unknown"
        try:
            with self.db.get_connection() as conn:
                query = conn.cursor()
                query.execute("SELECT username FROM users WHERE user_id = ?", (user_id,))
                result = query.fetchone()
                if result:
                    username = result[0]
        except Exception as e:
            logger.warning(f"Could not fetch username for user_id = {user_id}: {e}")

        # 4. Generate workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{order_number} Tracking"

        # 5. Write headers
        headers = [
                "Created By (admin)",
                "Created At",
                "Operator",
                "Company ID", 
                "Board ID", 
                "Serial Number", 
                "pass/fail", 
                "pass/fail timestamp", 
                "Failure Explanation (only if failed)", 
                "Repair Explanation (only if fixed)"
            ]
        ws.append(headers)

        # 6. styling
        for col in ws[1]:
            col.font = Font(bold=True)
            col.alignment = Alignment(horizontal="center", vertical="center")

        # 7. data alignment
        serials = self._generate_serial_numbers(prefix=serial_prefix, start=serial_start, count=serial_count)

        # 8. meta data
        status_value = "Pending"
        created_at = datetime.now().strftime("%b %d, %Y %I:%M %p")

        # 9. write serial rows
        for sn in serials:
            row_data = [
                created_by,
                created_at,
                username,                              # User ID
                company_id,                           # Company ID
                board_name if board_id else None,      # Board ID (None if not specified)
                sn,                                   # Serial Number
                status_value,                         # Pass/Fail
                None,                                 # timestamp
                None,                                 # fail explanation
                None                                  # fix explanation  
            ]
            ws.append(row_data)

        # 10. Format explanations
        for col_idx in (9,10):
            for row in range(2, len(serials)+ 2):
                ws.cell(row=row, column=col_idx).alignment = Alignment(wrap_text=True)

        # 11. Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            # Set minimum width and add padding
            ws.column_dimensions[col_letter].width = max(max_length + 2, 10)

        # 12. Save file
        wb.save(file_path)

        # 13. Register order in DB
        self.db.add_order(order_number, company_id, board_id, file_path, created_by)

        return file_path, len(serials)

    def is_order_ready_for_confirmation(self, file_path: str) -> bool:
        """Return True if every data row in the XLSX file has a passing pass/fail value
        and a non-empty pass/fail timestamp. This indicates the order is ready for admin confirmation.

        Heuristics:
        - Expects the header columns created by create_order_file. Finds the "pass/fail" and "pass/fail timestamp"
          columns by name.
        - A cell is considered a passing value if its string lower() is in ('pass','true','1','yes').
        - All data rows must match the above and have a non-empty timestamp.
        """
        if not file_path or not os.path.exists(file_path):
            return False

        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active

            # Find header indices
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            # Normalize header names
            header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers) if h is not None}
            pf_key = None
            ts_key = None
            for k in header_map:
                if 'pass' in k and 'fail' in k and 'timestamp' not in k:
                    # unlikely, but avoid
                    pass
            # find pass/fail header
            for name in ('pass/fail', 'passfail', 'pass_fail'):
                if name in header_map:
                    pf_key = header_map[name]
                    break
            # find timestamp header
            for name in ('pass/fail timestamp', 'pass/fail timestamp', 'pass_fail timestamp', 'pass/fail timestamp'):
                if name in header_map:
                    ts_key = header_map[name]
                    break

            # Fallback: try simple lookups by partial match
            if pf_key is None:
                for k in header_map:
                    if 'pass' in k and 'fail' in k and 'timestamp' not in k:
                        pf_key = header_map[k]
                        break
                    if 'pass' in k and 'fail' not in k and 'timestamp' not in k:
                        pf_key = header_map[k]
                        break

            if ts_key is None:
                for k in header_map:
                    if 'timestamp' in k:
                        ts_key = header_map[k]
                        break

            if pf_key is None or ts_key is None:
                # Unable to find relevant columns — not ready
                return False

            pass_values = {'pass', 'true', '1', 'yes'}

            # Iterate rows from row 2 onward
            for row in ws.iter_rows(min_row=2, values_only=True):
                # row is a tuple aligned with headers
                try:
                    pf_val = row[pf_key]
                except Exception:
                    return False

                # Normalize pf_val
                if pf_val is None:
                    return False
                pf_str = str(pf_val).strip().lower()
                if pf_str not in pass_values:
                    return False

                # Check timestamp
                try:
                    ts_val = row[ts_key]
                except Exception:
                    return False
                if ts_val is None or (isinstance(ts_val, str) and ts_val.strip() == ""):
                    return False

            return True
        except Exception:
            return False