# tests/test_xlsx_manager.py
import unittest
import tempfile
import os
import sys
import shutil
import random
from datetime import datetime

# Add project root to Python path
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

from openpyxl import load_workbook
from managers.db_manager import DatabaseManager
from managers.xlsx_manager import XLSXManager


class TestXLSXManager(unittest.TestCase):
    def setUp(self):
        # Temp dir for database + XLSX files
        self.test_dir = tempfile.mkdtemp()

        # Database in temp dir
        self.db_path = os.path.join(self.test_dir, "db")
        self.db = DatabaseManager(db_path=self.db_path, db_name="test.db")

        # Add test user
        self.db.add_user("test_user", "password123", role="admin")
        self.user_id = self.db.authenticate_user("test_user", "password123")[0]
        
        # Add test company (point its client_path inside temp dir)
        self.company_storage = os.path.join(self.test_dir, "TestCompany")
        self.db.add_company("Test Company", self.company_storage)
        self.company_id = self.db.get_companies()[0][0]

        # Add a test board for testing board_id functionality
        self.db.add_board(self.company_id, "Test Board")
        self.board_id = self.db.get_boards_by_company(self.company_id)[0][0]

        # XLSX manager
        self.xlsx_mgr = XLSXManager(self.db)

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def test_create_order_file(self):
        """Test XLSX file creation and DB order logging"""

        order_number = "ORD12345"
        pass_fail = random.choice([True, False])  # Fixed: random.choice instead of random(bool)
        pass_fail_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if pass_fail == False:
            failure_explanation = "it failed because xxxx"
            fix_explanation = "it was fixed by replaced x, y, and z. Retested and received results a, b, c"
        else:
            failure_explanation = ""
            fix_explanation = ""

        # Create order file with 10 serials (using correct parameter names from xlsx_manager.py)
        file_path, count = self.xlsx_mgr.create_order_file(
            order_number=order_number,
            created_by=self.user_id,
            user_id=self.user_id,
            company_id=self.company_id,
            pass_fail=pass_fail,
            pass_fail_timestamp=pass_fail_timestamp,
            failure_explanation=failure_explanation,
            fix_explanation=fix_explanation,
            board_id=None,
            serial_prefix="CUSTID-ORDNO-",
            serial_start=1,
            serial_count=10
        )

        # 1. File should exist
        self.assertTrue(os.path.exists(file_path))
        self.assertEqual(count, 10)  # Should return count of serials created

        # 2. Workbook should have correct headers
        wb = load_workbook(file_path)
        ws = wb.active

        # Check actual headers from xlsx_manager.py
        expected_headers = ["User ID", "Company ID", "Board ID", "Serial Number", 
                          "pass/fail", "pass/fail timestamp", "if failed explanation", 
                          "if fixed explanation"]
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, expected_headers)

        # Check row count (1 header + 10 data rows)
        self.assertEqual(ws.max_row, 11)

        # 3. Serial numbers should be CUSTID-ORDNO-00001 ... CUSTID-ORDNO-00010
        # Serial numbers are now correctly in column 4 (Serial Number column)
        serials = [ws.cell(row=i, column=4).value for i in range(2, 12)]
        self.assertEqual(serials[0], "CUSTID-ORDNO-00001")
        self.assertEqual(serials[-1], "CUSTID-ORDNO-00010")

        # 4. Check data structure matches headers properly
        # Data structure: [user_id, company_id, board_id, serial_number, pass_fail, timestamp, failure_explanation, fix_explanation]
        row_2_data = [ws.cell(row=2, column=i).value for i in range(1, 9)]  # All 8 columns
        
        self.assertEqual(row_2_data[0], self.user_id)  # User ID in column 1
        self.assertEqual(row_2_data[1], self.company_id)  # Company ID in column 2
        self.assertIsNone(row_2_data[2])  # Board ID in column 3 (None since board_id=None)
        self.assertEqual(row_2_data[3], "CUSTID-ORDNO-00001")  # Serial Number in column 4
        
        # Newly-created order files should start with status 'Pending' and no timestamp/explanations
        self.assertEqual(row_2_data[4], "Pending")  # pass/fail in column 5
        self.assertIsNone(row_2_data[5])  # Timestamp should be empty for new files
        self.assertIsNone(row_2_data[6])  # No failure explanation for new files
        self.assertIsNone(row_2_data[7])  # No fix explanation for new files

        # 5. Orders table in DB should have an entry
        orders = self.db.get_orders(self.company_id)
        self.assertEqual(len(orders), 1)
        self.assertEqual(orders[0][1], order_number)  # order_number

    def test_create_order_file_custom_parameters(self):
        """Test XLSX file creation with custom serial prefix and count"""
        
        order_number = "ORD67890"
        pass_fail_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        file_path, count = self.xlsx_mgr.create_order_file(
            order_number=order_number,
            created_by=self.user_id,
            user_id=self.user_id,
            company_id=self.company_id,
            pass_fail=True,
            pass_fail_timestamp=pass_fail_timestamp,
            failure_explanation="",
            fix_explanation="",
            board_id=self.board_id,  # Use the board_id we created in setUp
            serial_prefix="TEST-",
            serial_start=5,
            serial_count=3
        )
        
        # Check file exists and count is correct
        self.assertTrue(os.path.exists(file_path))
        self.assertEqual(count, 3)
        
        # Check serial numbers are correct (they're now in column 4 - Serial Number)
        wb = load_workbook(file_path)
        ws = wb.active
        
        serials = [ws.cell(row=i, column=4).value for i in range(2, 5)]
        self.assertEqual(serials, ["TEST-00005", "TEST-00006", "TEST-00007"])

    def test_generate_serial_numbers(self):
        """Test the serial number generation method"""
        
        # Test default parameters
        serials = self.xlsx_mgr._generate_serial_numbers()
        self.assertEqual(len(serials), 100)
        self.assertEqual(serials[0], "CUSTID-ORDNO-00001")
        self.assertEqual(serials[-1], "CUSTID-ORDNO-00100")
        
        # Test custom parameters
        custom_serials = self.xlsx_mgr._generate_serial_numbers(
            prefix="CUSTOM-", start=10, count=5
        )
        expected = ["CUSTOM-00010", "CUSTOM-00011", "CUSTOM-00012", 
                   "CUSTOM-00013", "CUSTOM-00014"]
        self.assertEqual(custom_serials, expected)

    def test_invalid_company_id(self):
        """Test that creating an order for a non-existent company raises ValueError"""
        pass_fail_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        with self.assertRaises(ValueError) as context:
            self.xlsx_mgr.create_order_file(
                order_number="ORD999",
                created_by=self.user_id,
                user_id=self.user_id,
                company_id=9999,  # Nonexistent company
                pass_fail=True,
                pass_fail_timestamp=pass_fail_timestamp,
                failure_explanation="",
                fix_explanation=""
            )
        
        self.assertIn("Company 9999 not found", str(context.exception))

    def test_file_formatting(self):
        """Test that the XLSX file has proper formatting"""
        
        order_number = "FORMAT_TEST"
        pass_fail_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        file_path, _ = self.xlsx_mgr.create_order_file(
            order_number=order_number,
            created_by=self.user_id,
            user_id=self.user_id,
            company_id=self.company_id,
            pass_fail=True,
            pass_fail_timestamp=pass_fail_timestamp,
            failure_explanation="",
            fix_explanation="",
            serial_count=2
        )
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Check that headers are bold and centered
        for col in ws[1]:
            self.assertTrue(col.font.bold)
            self.assertEqual(col.alignment.horizontal, "center")
            self.assertEqual(col.alignment.vertical, "center")
        
        # Check that failure explanation columns have wrap_text
        # Column 7 = "if failed explanation", Column 8 = "if fixed explanation"
        for row in range(2, 4):  # Check first 2 data rows
            self.assertTrue(ws.cell(row=row, column=7).alignment.wrap_text)
            self.assertTrue(ws.cell(row=row, column=8).alignment.wrap_text)

    def test_pass_fail_scenarios(self):
        """Test both passing and failing scenarios specifically"""
        
        # Test passing scenario
        order_number_pass = "ORD_PASS"
        pass_fail_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        file_path_pass, _ = self.xlsx_mgr.create_order_file(
            order_number=order_number_pass,
            created_by=self.user_id,
            user_id=self.user_id,
            company_id=self.company_id,
            pass_fail=True,  # PASS
            pass_fail_timestamp=pass_fail_timestamp,
            failure_explanation="This should not appear",
            fix_explanation="This should not appear",
            serial_count=1
        )
        
        wb_pass = load_workbook(file_path_pass)
        ws_pass = wb_pass.active
        
        # Newly-created files always start as Pending regardless of pass_fail flag
        self.assertEqual(ws_pass.cell(row=2, column=5).value, "Pending")
        self.assertIsNone(ws_pass.cell(row=2, column=7).value)  # failure explanation None
        self.assertIsNone(ws_pass.cell(row=2, column=8).value)  # fix explanation None
        
        # Test failing scenario
        order_number_fail = "ORD_FAIL"
        failure_explanation = "Component X failed voltage test"
        fix_explanation = "Replaced component X and retested successfully"
        
        file_path_fail, _ = self.xlsx_mgr.create_order_file(
            order_number=order_number_fail,
            created_by=self.user_id,
            user_id=self.user_id,
            company_id=self.company_id,
            pass_fail=False,  # FAIL
            pass_fail_timestamp=pass_fail_timestamp,
            failure_explanation=failure_explanation,
            fix_explanation=fix_explanation,
            serial_count=1
        )
        
        wb_fail = load_workbook(file_path_fail)
        ws_fail = wb_fail.active
        
        # Newly-created files always start as Pending regardless of pass_fail flag
        self.assertEqual(ws_fail.cell(row=2, column=5).value, "Pending")
        self.assertIsNone(ws_fail.cell(row=2, column=7).value)
        self.assertIsNone(ws_fail.cell(row=2, column=8).value)


if __name__ == "__main__":
    unittest.main(verbosity=2)